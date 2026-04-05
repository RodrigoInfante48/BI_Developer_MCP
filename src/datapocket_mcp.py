#!/usr/bin/env python3
"""
DataPocket MCP Server v2 — BI Agent for Everyone.

Turn Claude into a personal Business Intelligence agent that:
1. Ingests data from CSV / TSV / JSON / JSONL / Excel (XLSX)
2. Transforms it with Python operations
3. Loads it into PostgreSQL
4. Generates mobile-first HTML dashboards
5. Exports data in multiple formats

Stack: Python + PostgreSQL + Power BI (local) — zero licensing costs.
Product by DailyDuty / Instituto para el Desarrollo Diario.
Version: 2.0.0
"""

import json
import os
import csv
import io
import hashlib
from typing import Optional, List, Dict, Any, Tuple
from enum import Enum
from datetime import datetime

from pydantic import BaseModel, Field, field_validator, ConfigDict
from mcp.server.fastmcp import FastMCP, Context

try:
    import chardet
    _CHARDET_AVAILABLE = True
except ImportError:
    _CHARDET_AVAILABLE = False

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# ──────────────────────────────────────────────
# Server Initialization
# ──────────────────────────────────────────────
mcp = FastMCP("datapocket_mcp")

DEFAULT_SCHEMA = "public"
MAX_PREVIEW_ROWS = 50
DASHBOARD_OUTPUT_DIR = os.environ.get("DATAPOCKET_DASHBOARDS_DIR", "./dashboards")


# ──────────────────────────────────────────────
# Custom Exceptions
# ──────────────────────────────────────────────

class DataPocketError(Exception):
    """Base error for DataPocket MCP."""
    pass


class ParseError(DataPocketError):
    """Error parsing input data."""
    pass


class ValidationError(DataPocketError):
    """Error validating data structure."""
    pass


class TransformError(DataPocketError):
    """Error during data transformation."""
    pass


# ──────────────────────────────────────────────
# Enums & Shared Models
# ──────────────────────────────────────────────

class ResponseFormat(str, Enum):
    MARKDOWN = "markdown"
    JSON = "json"


class ChartType(str, Enum):
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    KPI = "kpi"
    TABLE = "table"
    AREA = "area"
    DONUT = "donut"


class DataSourceType(str, Enum):
    CSV = "csv"
    JSON_DATA = "json"
    JSONL = "jsonl"
    TSV = "tsv"
    XLSX = "xlsx"
    AUTO = "auto"


class ExportFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"
    SQL = "sql"
    MARKDOWN_TABLE = "markdown"


# ──────────────────────────────────────────────
# Encoding & Format Detection Utilities
# ──────────────────────────────────────────────

def _detect_and_decode(raw_data) -> str:
    """Detect encoding and decode bytes to string.

    Args:
        raw_data: Either a str (returned as-is) or bytes to decode.

    Returns:
        str: Decoded text using detected or fallback encoding.
    """
    if isinstance(raw_data, str):
        return raw_data
    if _CHARDET_AVAILABLE:
        detected = chardet.detect(raw_data)
        encoding = detected.get("encoding", "utf-8") or "utf-8"
    else:
        encoding = "utf-8"
    # Fallback chain
    for enc in [encoding, "utf-8", "latin-1"]:
        try:
            return raw_data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw_data.decode("utf-8", errors="replace")


def _auto_detect_format(data: str) -> DataSourceType:
    """Detect the format of raw data string.

    Args:
        data: Raw data string to inspect.

    Returns:
        DataSourceType: Detected format (csv, json, jsonl, or tsv).
    """
    stripped = data.strip()
    if not stripped:
        return DataSourceType.CSV

    # JSON array
    if stripped.startswith("["):
        return DataSourceType.JSON_DATA

    # JSONL — each non-empty line starts with {
    lines = [l.strip() for l in stripped.splitlines() if l.strip()]
    if lines and all(l.startswith("{") for l in lines[:5]):
        return DataSourceType.JSONL

    # TSV — tab-separated
    first_line = lines[0] if lines else ""
    if "\t" in first_line:
        tab_count = first_line.count("\t")
        comma_count = first_line.count(",")
        if tab_count >= comma_count:
            return DataSourceType.TSV

    return DataSourceType.CSV


# ──────────────────────────────────────────────
# Shared Utilities
# ──────────────────────────────────────────────

def _get_pg_connection_string() -> str:
    """Build PostgreSQL connection string from environment variables.

    Returns:
        str: PostgreSQL connection URL.
    """
    host = os.environ.get("DATAPOCKET_PG_HOST", "localhost")
    port = os.environ.get("DATAPOCKET_PG_PORT", "5432")
    db = os.environ.get("DATAPOCKET_PG_DB", "datapocket")
    user = os.environ.get("DATAPOCKET_PG_USER", "datapocket_user")
    password = os.environ.get("DATAPOCKET_PG_PASSWORD", "")
    return f"postgresql://{user}:{password}@{host}:{port}/{db}"


def _sanitize_table_name(name: str) -> str:
    """Sanitize a string to be a valid PostgreSQL table/column name.

    Args:
        name: Raw name to sanitize.

    Returns:
        str: Sanitized identifier, max 63 chars.
    """
    sanitized = "".join(c if c.isalnum() or c == "_" else "_" for c in name.lower())
    if not sanitized or sanitized[0].isdigit():
        sanitized = f"t_{sanitized}"
    return sanitized[:63]


def _infer_pg_type(values: List[Any]) -> str:
    """Infer PostgreSQL column type from a sample of values.

    Args:
        values: List of raw values to inspect.

    Returns:
        str: PostgreSQL type string (BIGINT, NUMERIC, DATE, TIMESTAMP, BOOLEAN, VARCHAR, TEXT).
    """
    sample = [v for v in values if v is not None and str(v).strip() != ""][:100]
    if not sample:
        return "TEXT"

    try:
        [int(str(v).replace(",", "")) for v in sample]
        return "BIGINT"
    except (ValueError, TypeError):
        pass

    try:
        [float(str(v).replace(",", "")) for v in sample]
        return "NUMERIC"
    except (ValueError, TypeError):
        pass

    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]
    for fmt in date_formats:
        try:
            [datetime.strptime(str(v).strip(), fmt) for v in sample]
            return "TIMESTAMP" if " " in fmt else "DATE"
        except (ValueError, TypeError):
            continue

    bool_values = {"true", "false", "1", "0", "yes", "no", "si", "sí"}
    if all(str(v).strip().lower() in bool_values for v in sample):
        return "BOOLEAN"

    max_len = max(len(str(v)) for v in sample) if sample else 255
    if max_len <= 50:
        return "VARCHAR(100)"
    elif max_len <= 255:
        return "VARCHAR(500)"
    return "TEXT"


def _smart_format(value: float, prefix: str = "", suffix: str = "") -> str:
    """Format a number with K/M/B suffix for readability.

    Args:
        value: Numeric value to format.
        prefix: String to prepend (e.g. '$').
        suffix: String to append (e.g. '%').

    Returns:
        str: Human-readable formatted number.
    """
    if abs(value) >= 1_000_000_000:
        return f"{prefix}{value / 1_000_000_000:.1f}B{suffix}"
    elif abs(value) >= 1_000_000:
        return f"{prefix}{value / 1_000_000:.1f}M{suffix}"
    elif abs(value) >= 1_000:
        return f"{prefix}{value / 1_000:.1f}K{suffix}"
    elif value == int(value):
        return f"{prefix}{int(value)}{suffix}"
    return f"{prefix}{value:.2f}{suffix}"


def _parse_csv_data(csv_text: str) -> Tuple[List[str], List[List[str]]]:
    """Parse CSV text into headers and rows.

    Args:
        csv_text: Raw CSV string with headers in the first row.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)

    Raises:
        ParseError: If CSV is empty or has no headers.
    """
    csv_text = _detect_and_decode(csv_text)
    reader = csv.reader(io.StringIO(csv_text))
    rows = list(reader)
    if not rows:
        raise ParseError("No se encontraron headers en el CSV. Sugerencia: Verifica que la primera fila contenga los nombres de columna.")
    headers = [_sanitize_table_name(h.strip()) for h in rows[0] if h.strip()]
    if not headers:
        raise ParseError("La primera fila del CSV está vacía. Sugerencia: Agrega headers en la primera fila.")
    data_rows = [r for r in rows[1:] if any(str(v).strip() for v in r)]
    return headers, data_rows


def _parse_tsv_data(tsv_text: str) -> Tuple[List[str], List[List[str]]]:
    """Parse TSV (Tab-Separated Values) text into headers and rows.

    Args:
        tsv_text: Raw TSV string with headers in the first row.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)

    Raises:
        ParseError: If TSV is empty or has no headers.
    """
    tsv_text = _detect_and_decode(tsv_text)
    reader = csv.reader(io.StringIO(tsv_text), delimiter="\t")
    rows = list(reader)
    if not rows:
        raise ParseError("TSV vacío. Sugerencia: Verifica que el archivo tenga contenido con tabs como separador.")
    headers = [_sanitize_table_name(h.strip()) for h in rows[0] if h.strip()]
    data_rows = [r for r in rows[1:] if any(str(v).strip() for v in r)]
    return headers, data_rows


def _parse_jsonl_data(jsonl_text: str) -> Tuple[List[str], List[List[str]]]:
    """Parse JSON Lines text into headers and rows.

    Args:
        jsonl_text: Raw JSONL string — one JSON object per line.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)

    Raises:
        ParseError: If no valid JSON lines are found.
    """
    jsonl_text = _detect_and_decode(jsonl_text)
    records = []
    for i, line in enumerate(jsonl_text.splitlines()):
        line = line.strip()
        if not line:
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError as e:
            raise ParseError(f"Línea {i+1} no es JSON válido: {e}. Sugerencia: Cada línea debe ser un objeto JSON completo.")
    if not records:
        raise ParseError("No se encontraron líneas JSON válidas. Sugerencia: Verifica que el archivo JSONL tenga al least una línea de datos.")
    headers = [_sanitize_table_name(k) for k in records[0].keys()]
    data_rows = [[str(r.get(k, "")) for k in records[0].keys()] for r in records]
    return headers, data_rows


def _parse_json_data(json_text: str) -> Tuple[List[str], List[List[str]]]:
    """Parse JSON array text into headers and rows.

    Args:
        json_text: Raw JSON string — array of objects.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)

    Raises:
        ParseError: If JSON is invalid or not an array.
    """
    json_text = _detect_and_decode(json_text)
    try:
        records = json.loads(json_text)
    except json.JSONDecodeError as e:
        raise ParseError(f"JSON inválido: {e}. Sugerencia: Verifica que el JSON sea un array de objetos [{'{'}...{'}'}].")
    if not isinstance(records, list):
        raise ParseError("El JSON debe ser un array de objetos. Sugerencia: Envuelve los datos en corchetes: [{...}].")
    if not records:
        raise ParseError("JSON array vacío. Sugerencia: Agrega al menos un objeto al array.")
    headers = [_sanitize_table_name(k) for k in records[0].keys()]
    data_rows = [[str(r.get(k, "")) for k in records[0].keys()] for r in records]
    return headers, data_rows


def _parse_xlsx_data(xlsx_path_or_b64: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[List[str]]]:
    """Parse an XLSX file (by path) into headers and rows.

    Args:
        xlsx_path_or_b64: Path to the XLSX file.
        sheet_name: Name of the sheet to read. Defaults to the first sheet.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)

    Raises:
        ParseError: If openpyxl is not installed or the file cannot be read.
    """
    if not _OPENPYXL_AVAILABLE:
        raise ParseError("openpyxl no está instalado. Sugerencia: Ejecuta 'pip install openpyxl'.")
    try:
        wb = openpyxl.load_workbook(xlsx_path_or_b64, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ParseError("El archivo Excel está vacío.")
        headers = [_sanitize_table_name(str(h).strip()) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = [[str(v) if v is not None else "" for v in row] for row in rows[1:]]
        return headers, data_rows
    except ParseError:
        raise
    except Exception as e:
        raise ParseError(f"No se pudo leer el archivo Excel: {e}. Sugerencia: Verifica que el path sea correcto y que sea un archivo .xlsx válido.")


def _parse_any(data: str, source_type: DataSourceType) -> Tuple[List[str], List[List[str]]]:
    """Unified parser that dispatches to the correct format parser.

    Args:
        data: Raw data string.
        source_type: Explicit format or AUTO for auto-detection.

    Returns:
        Tuple[List[str], List[List[str]]]: (headers, data_rows)
    """
    if source_type == DataSourceType.AUTO:
        source_type = _auto_detect_format(data)
    if source_type == DataSourceType.CSV:
        return _parse_csv_data(data)
    elif source_type == DataSourceType.TSV:
        return _parse_tsv_data(data)
    elif source_type == DataSourceType.JSON_DATA:
        return _parse_json_data(data)
    elif source_type == DataSourceType.JSONL:
        return _parse_jsonl_data(data)
    elif source_type == DataSourceType.XLSX:
        return _parse_xlsx_data(data)
    else:
        return _parse_csv_data(data)


def _generate_create_table_sql(
    table_name: str,
    headers: List[str],
    data_rows: List[List[str]],
    schema: str = DEFAULT_SCHEMA
) -> str:
    """Generate CREATE TABLE SQL statement with inferred PostgreSQL types.

    Args:
        table_name: Target table name (already sanitized).
        headers: List of column names.
        data_rows: List of data rows.
        schema: PostgreSQL schema name.

    Returns:
        str: DROP + CREATE TABLE SQL statement.
    """
    columns = []
    for i, header in enumerate(headers):
        col_values = [row[i] if i < len(row) else None for row in data_rows]
        pg_type = _infer_pg_type(col_values)
        columns.append(f'    "{header}" {pg_type}')
    columns_sql = ",\n".join(columns)
    return f"""DROP TABLE IF EXISTS {schema}."{table_name}" CASCADE;
CREATE TABLE {schema}."{table_name}" (
    id SERIAL PRIMARY KEY,
{columns_sql}
);"""


def _generate_insert_sql(
    table_name: str,
    headers: List[str],
    data_rows: List[List[str]],
    schema: str = DEFAULT_SCHEMA
) -> str:
    """Generate batched INSERT SQL statements.

    Args:
        table_name: Target table name (already sanitized).
        headers: List of column names.
        data_rows: List of data rows.
        schema: PostgreSQL schema name.

    Returns:
        str: One or more INSERT statements batched in groups of 100.
    """
    if not data_rows:
        return "-- No data rows to insert"
    cols = ", ".join(f'"{h}"' for h in headers)
    values_list = []
    for row in data_rows:
        vals = []
        for v in row:
            if v is None or str(v).strip() == "":
                vals.append("NULL")
            else:
                escaped = str(v).replace("'", "''")
                vals.append(f"'{escaped}'")
        while len(vals) < len(headers):
            vals.append("NULL")
        values_list.append(f"({', '.join(vals[:len(headers)])})")
    batches = []
    for i in range(0, len(values_list), 100):
        batch = values_list[i:i + 100]
        batches.append(
            f'INSERT INTO {schema}."{table_name}" ({cols})\nVALUES\n' +
            ",\n".join(batch) + ";"
        )
    return "\n\n".join(batches)


# ──────────────────────────────────────────────
# Dashboard HTML Generator (Mobile-First)
# ──────────────────────────────────────────────

def _generate_dashboard_html(
    title: str,
    charts: List[Dict[str, Any]],
    theme: str = "dark",
    data_timestamp: Optional[str] = None,
) -> str:
    """Generate a mobile-first HTML dashboard with Chart.js.

    Args:
        title: Dashboard title displayed in the header.
        charts: List of chart/KPI/table config dicts.
        theme: 'dark' or 'light' color scheme.
        data_timestamp: Optional date string representing the data period.

    Returns:
        str: Complete self-contained HTML with embedded CSS and JS.
    """
    if theme == "light":
        bg = "#f8fafc"
        text = "#1e293b"
        card_bg = "#ffffff"
        border = "#e2e8f0"
    else:
        bg = "#0f172a"
        text = "#e2e8f0"
        card_bg = "#1e293b"
        border = "#334155"

    accent = "#38bdf8"
    accent2 = "#a78bfa"
    accent3 = "#34d399"

    charts_js = []
    charts_html = []

    # Separate KPI charts for grid layout
    kpi_charts = [c for c in charts if c.get("type") == "kpi"]
    other_charts = [c for c in charts if c.get("type") != "kpi"]

    # KPI grid
    if kpi_charts:
        kpi_items_html = []
        for kpi in kpi_charts:
            kpi_value = kpi.get("value", "0")
            kpi_subtitle = kpi.get("subtitle", "")
            kpi_trend = kpi.get("trend", "")
            trend_color = accent3 if "+" in str(kpi_trend) or "↑" in str(kpi_trend) else "#f87171"
            if not kpi_trend:
                trend_color = text
            kpi_items_html.append(f"""        <div class="card kpi-card">
            <div class="kpi-label">{kpi.get('title', '')}</div>
            <div class="kpi-value">{kpi_value}</div>
            <div class="kpi-subtitle">{kpi_subtitle}</div>
            <div class="kpi-trend" style="color:{trend_color}">{kpi_trend}</div>
        </div>""")

        n = len(kpi_charts)
        if n == 1:
            charts_html.append(kpi_items_html[0])
        else:
            charts_html.append(f'        <div class="kpi-grid">\n' + "\n".join(kpi_items_html) + "\n        </div>")

    for i, chart in enumerate(other_charts):
        chart_id = f"chart_{i}"
        chart_type = chart.get("type", "bar")
        chart_title = chart.get("title", f"Chart {i + 1}")
        labels = json.dumps(chart.get("labels", []))
        datasets = chart.get("datasets", [])

        if chart_type == "table":
            table_headers = chart.get("headers", [])
            table_rows = chart.get("rows", [])
            th_html = "".join(f"<th>{h}</th>" for h in table_headers)
            tr_html = ""
            for row in table_rows[:20]:
                td_html = "".join(f"<td>{v}</td>" for v in row)
                tr_html += f"<tr>{td_html}</tr>"
            charts_html.append(f"""        <div class="card table-card">
            <h3>{chart_title}</h3>
            <div class="table-scroll">
                <table><thead><tr>{th_html}</tr></thead><tbody>{tr_html}</tbody></table>
            </div>
        </div>""")
            continue

        colors = [accent, accent2, accent3, "#fb923c", "#f472b6", "#facc15"]
        ds_json_parts = []
        for j, ds in enumerate(datasets):
            color = colors[j % len(colors)]
            n_pts = len(ds.get("data", []))
            ds_entry = {
                "label": ds.get("label", f"Serie {j + 1}"),
                "data": ds.get("data", []),
                "backgroundColor": f"{color}88" if chart_type in ("bar", "area") else [f"{colors[k % len(colors)]}cc" for k in range(n_pts)],
                "borderColor": color,
                "borderWidth": 2,
                "borderRadius": 6 if chart_type == "bar" else 0,
                "tension": 0.4 if chart_type in ("line", "area") else 0,
                "fill": chart_type == "area",
                "pointRadius": 4 if chart_type == "line" else 0,
            }
            ds_json_parts.append(json.dumps(ds_entry))

        datasets_js = "[" + ",".join(ds_json_parts) + "]"
        ct = "doughnut" if chart_type == "donut" else ("line" if chart_type == "area" else chart_type)

        charts_html.append(f"""        <div class="card chart-card">
            <h3>{chart_title}</h3>
            <div class="chart-container">
                <canvas id="{chart_id}"></canvas>
            </div>
        </div>""")

        charts_js.append(f"""        new Chart(document.getElementById('{chart_id}'), {{
            type: '{ct}',
            data: {{ labels: {labels}, datasets: {datasets_js} }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{ display: {str(len(datasets) > 1).lower()}, labels: {{ color: '{text}', font: {{ size: 11 }} }} }}
                }},
                scales: '{ct}' === 'doughnut' || '{ct}' === 'pie' ? {{}} : {{
                    x: {{ ticks: {{ color: '{text}', font: {{ size: 10 }} }}, grid: {{ color: '{border}33' }} }},
                    y: {{ ticks: {{ color: '{text}', font: {{ size: 10 }} }}, grid: {{ color: '{border}33' }} }}
                }}
            }}
        }});""")

    charts_html_str = "\n".join(charts_html)
    charts_js_str = "\n".join(charts_js)
    gen_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    data_ts_line = f'<div class="timestamp">Datos: {data_timestamp}</div>' if data_timestamp else ""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
    <title>{title} — DataPocket</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&display=swap');
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'DM Sans', sans-serif;
            background: {bg};
            color: {text};
            min-height: 100vh;
            -webkit-font-smoothing: antialiased;
        }}
        .dashboard {{
            max-width: 480px;
            margin: 0 auto;
            padding: 16px;
            padding-bottom: 80px;
        }}
        .dash-header {{
            text-align: center;
            padding: 20px 0 16px;
        }}
        .dash-header h1 {{
            font-size: 1.3rem;
            font-weight: 700;
            letter-spacing: -0.02em;
        }}
        .dash-header .timestamp {{
            font-size: 0.7rem;
            opacity: 0.5;
            margin-top: 4px;
        }}
        .dash-header .brand {{
            font-size: 0.65rem;
            opacity: 0.35;
            margin-top: 2px;
            letter-spacing: 0.1em;
            text-transform: uppercase;
        }}
        .card {{
            background: {card_bg};
            border: 1px solid {border};
            border-radius: 16px;
            padding: 16px;
            margin-bottom: 12px;
            animation: fadeUp 0.4s ease both;
        }}
        .card h3 {{
            font-size: 0.8rem;
            font-weight: 500;
            opacity: 0.7;
            margin-bottom: 12px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 10px;
            margin-bottom: 12px;
        }}
        .kpi-grid .kpi-card {{ margin-bottom: 0; }}
        .kpi-card {{
            text-align: center;
            padding: 20px 16px;
        }}
        .kpi-label {{
            font-size: 0.75rem;
            opacity: 0.6;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            margin-bottom: 4px;
        }}
        .kpi-value {{
            font-size: 2rem;
            font-weight: 700;
            color: {accent};
            line-height: 1.1;
        }}
        .kpi-subtitle {{
            font-size: 0.7rem;
            opacity: 0.5;
            margin-top: 4px;
        }}
        .kpi-trend {{
            font-size: 0.8rem;
            font-weight: 600;
            margin-top: 6px;
        }}
        .chart-card .chart-container {{
            position: relative;
            height: 220px;
        }}
        .table-scroll {{
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.75rem;
        }}
        th, td {{
            padding: 8px 10px;
            text-align: left;
            border-bottom: 1px solid {border};
            white-space: nowrap;
        }}
        th {{
            font-weight: 600;
            opacity: 0.7;
            text-transform: uppercase;
            font-size: 0.65rem;
            letter-spacing: 0.05em;
        }}
        @keyframes fadeUp {{
            from {{ opacity: 0; transform: translateY(12px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        .card:nth-child(1) {{ animation-delay: 0s; }}
        .card:nth-child(2) {{ animation-delay: 0.05s; }}
        .card:nth-child(3) {{ animation-delay: 0.1s; }}
        .card:nth-child(4) {{ animation-delay: 0.15s; }}
        .card:nth-child(5) {{ animation-delay: 0.2s; }}
        .card:nth-child(6) {{ animation-delay: 0.25s; }}
        .refresh-hint {{
            text-align: center;
            font-size: 0.65rem;
            opacity: 0.3;
            padding: 12px;
        }}
        .bottom-bar {{
            position: fixed;
            bottom: 0;
            left: 50%;
            transform: translateX(-50%);
            width: 100%;
            max-width: 480px;
            background: {card_bg};
            border-top: 1px solid {border};
            padding: 10px 16px;
            display: flex;
            justify-content: space-around;
            z-index: 100;
        }}
        .bottom-bar button {{
            background: none;
            border: none;
            color: {text};
            opacity: 0.5;
            font-size: 0.7rem;
            cursor: pointer;
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 2px;
            min-height: 44px;
            justify-content: center;
        }}
        .bottom-bar button.active {{ opacity: 1; color: {accent}; }}
        .bottom-bar button svg {{ width: 20px; height: 20px; }}
    </style>
</head>
<body>
    <div class="dashboard">
        <div class="dash-header">
            <h1>{title}</h1>
            {data_ts_line}
            <div class="timestamp">Generado: {gen_timestamp}</div>
            <div class="brand">Powered by DataPocket</div>
        </div>

{charts_html_str}

        <div class="refresh-hint">Desliza hacia abajo para actualizar</div>
    </div>

    <div class="bottom-bar">
        <button class="active">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>
            Dashboard
        </button>
        <button onclick="location.reload()">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M1 4v6h6"/><path d="M23 20v-6h-6"/><path d="M20.49 9A9 9 0 0 0 5.64 5.64L1 10m22 4l-4.64 4.36A9 9 0 0 1 3.51 15"/></svg>
            Refresh
        </button>
        <button>
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.32 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z"/></svg>
            Config
        </button>
    </div>

    <script>
{charts_js_str}
    </script>
</body>
</html>"""


# ──────────────────────────────────────────────
# MCP Tools
# ──────────────────────────────────────────────

# ─── Tool 1: Ingest & Profile Data ───

class DataIngestInput(BaseModel):
    """Input for data ingestion and profiling."""
    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    data: str = Field(..., description="Raw data — CSV, TSV, JSON array, JSONL, or XLSX path")
    source_type: DataSourceType = Field(
        default=DataSourceType.AUTO,
        description="Format: 'csv', 'tsv', 'json', 'jsonl', 'xlsx', or 'auto' (default) for automatic detection"
    )
    table_name: str = Field(..., description="Target PostgreSQL table name (will be sanitized)", min_length=1, max_length=63)
    schema_name: str = Field(default=DEFAULT_SCHEMA, description="PostgreSQL schema name")

    @field_validator("table_name")
    @classmethod
    def validate_table_name(cls, v: str) -> str:
        """Sanitize table name to a valid PostgreSQL identifier."""
        return _sanitize_table_name(v)


@mcp.tool(
    name="datapocket_ingest_data",
    annotations={
        "title": "Ingest & Profile Data",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def datapocket_ingest_data(params: DataIngestInput) -> str:
    """Ingest raw data (CSV/TSV/JSON/JSONL/XLSX), profile it, and generate SQL for PostgreSQL.

    Analyzes column types, computes null statistics, and generates CREATE TABLE
    and batched INSERT statements ready to copy into any PostgreSQL client.

    Args:
        params (DataIngestInput): Validated input containing:
            - data (str): Raw data string or XLSX file path
            - source_type (DataSourceType): 'csv', 'tsv', 'json', 'jsonl', 'xlsx', or 'auto'
            - table_name (str): Target PostgreSQL table name
            - schema_name (str): Target schema (default: 'public')

    Returns:
        str: Markdown report with data profile + SQL statements ready to execute.
    """
    try:
        headers, data_rows = _parse_any(params.data, params.source_type)

        if not headers or not data_rows:
            return "❌ Error (ValidationError): No se encontraron datos después del parsing. Sugerencia: Verifica que el archivo tenga al menos una fila de datos."

        create_sql = _generate_create_table_sql(params.table_name, headers, data_rows, params.schema_name)
        insert_sql = _generate_insert_sql(params.table_name, headers, data_rows, params.schema_name)

        total_rows = len(data_rows)
        total_cols = len(headers)
        null_counts = {}
        for i, h in enumerate(headers):
            nulls = sum(1 for row in data_rows if i >= len(row) or row[i] is None or str(row[i]).strip() == "")
            null_counts[h] = nulls

        preview_rows = data_rows[:5]
        preview_table = "| " + " | ".join(headers) + " |\n"
        preview_table += "| " + " | ".join(["---"] * len(headers)) + " |\n"
        for row in preview_rows:
            preview_table += "| " + " | ".join(str(v)[:30] for v in row[:len(headers)]) + " |\n"

        report = f"""# 📊 DataPocket — Data Profile

## Summary
- **Table**: `{params.schema_name}.{params.table_name}`
- **Rows**: {total_rows:,}
- **Columns**: {total_cols}
- **Format detected**: `{params.source_type.value}`

## Column Analysis
| Column | Inferred Type | Nulls | Null % |
| --- | --- | --- | --- |
"""
        for i, h in enumerate(headers):
            col_values = [row[i] if i < len(row) else None for row in data_rows]
            pg_type = _infer_pg_type(col_values)
            null_pct = (null_counts[h] / total_rows * 100) if total_rows > 0 else 0
            report += f"| `{h}` | {pg_type} | {null_counts[h]} | {null_pct:.1f}% |\n"

        report += f"""
## Data Preview (first 5 rows)
{preview_table}

## SQL — CREATE TABLE
```sql
{create_sql}
```

## SQL — INSERT DATA ({total_rows} rows)
```sql
{insert_sql}
```

---
> ✅ Copy the SQL above and execute in your PostgreSQL client (psql, DBeaver, pgAdmin).
> Next step: Use `datapocket_generate_dashboard` to visualize this data.
"""
        return report

    except (ParseError, ValidationError) as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}"
    except Exception as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}. Sugerencia: Verifica el formato de los datos de entrada."


# ─── Tool 2: Transform Data ───

def _parse_op_parts(op: str) -> List[str]:
    """Parse operation string supporting both '|' and ':' separators.

    Args:
        op: Operation string like 'filter|col|>|100' or 'filter:col:>:100'.

    Returns:
        List[str]: Parts split by the dominant separator.
    """
    if "|" in op:
        return op.split("|")
    return op.split(":")


class TransformInput(BaseModel):
    """Input for Python data transformation."""
    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    data: str = Field(..., description="CSV data to transform (with headers)")
    operations: List[str] = Field(
        ...,
        description=(
            "Transformations to apply (use '|' as separator). Supported: "
            "'drop_nulls', 'dedup', 'sort|COLUMN|asc|desc', 'rename|OLD|NEW', "
            "'filter|COLUMN|OPERATOR|VALUE', 'lowercase|COLUMN', 'to_numeric|COLUMN', "
            "'aggregate|GROUP_COL|AGG_COL|sum|avg|count|min|max', "
            "'top|N|COLUMN', 'date_extract|COLUMN|year|month|day|weekday'"
        ),
        min_length=1,
        max_length=20
    )
    output_table_name: Optional[str] = Field(default=None, description="If provided, generates SQL for the transformed data")


@mcp.tool(
    name="datapocket_transform",
    annotations={
        "title": "Transform Data with Python",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": False,
        "openWorldHint": False
    }
)
async def datapocket_transform(params: TransformInput) -> str:
    """Transform CSV data with pandas-style operations and generate Python + SQL code.

    Applies a sequence of named operations to the input CSV and returns
    the transformed data, reproducible Python code, and optional SQL.

    Separator: use '|' (preferred) or ':' (legacy). Example: 'filter|ventas|>|1000'

    Args:
        params (TransformInput): Validated input containing:
            - data (str): CSV data with headers
            - operations (List[str]): Transformations to apply
            - output_table_name (Optional[str]): Target table for SQL generation

    Returns:
        str: Markdown with transformed data, Python code, and optional SQL.
    """
    try:
        headers, data_rows = _parse_csv_data(params.data)
        python_lines = [
            "import pandas as pd",
            "",
            "# Load data",
            "df = pd.read_csv('input.csv')",
            ""
        ]

        for op in params.operations:
            parts = _parse_op_parts(op)
            action = parts[0].lower().strip()

            if action == "drop_nulls":
                data_rows = [r for r in data_rows if all(str(v).strip() for v in r)]
                python_lines.append("df = df.dropna()")

            elif action == "dedup":
                seen = set()
                unique_rows = []
                for r in data_rows:
                    key = tuple(r)
                    if key not in seen:
                        seen.add(key)
                        unique_rows.append(r)
                data_rows = unique_rows
                python_lines.append("df = df.drop_duplicates()")

            elif action == "sort" and len(parts) >= 2:
                col = parts[1].strip()
                direction = parts[2].strip() if len(parts) > 2 else "asc"
                col_san = _sanitize_table_name(col)
                if col_san in headers:
                    col_idx = headers.index(col_san)
                    reverse = direction.lower() == "desc"
                    try:
                        data_rows.sort(key=lambda r: float(str(r[col_idx]).replace(",", "")) if col_idx < len(r) else 0, reverse=reverse)
                    except ValueError:
                        data_rows.sort(key=lambda r: str(r[col_idx]) if col_idx < len(r) else "", reverse=reverse)
                    python_lines.append(f"df = df.sort_values('{col_san}', ascending={not reverse})")

            elif action == "rename" and len(parts) >= 3:
                old_name = _sanitize_table_name(parts[1].strip())
                new_name = _sanitize_table_name(parts[2].strip())
                if old_name in headers:
                    idx = headers.index(old_name)
                    headers[idx] = new_name
                    python_lines.append(f"df = df.rename(columns={{'{old_name}': '{new_name}'}})")

            elif action == "filter" and len(parts) >= 4:
                col = _sanitize_table_name(parts[1].strip())
                operator = parts[2].strip()
                value = parts[3].strip()
                if col in headers:
                    col_idx = headers.index(col)
                    filtered = []
                    for r in data_rows:
                        cell = r[col_idx] if col_idx < len(r) else ""
                        try:
                            cell_num = float(str(cell).replace(",", ""))
                            val_num = float(value)
                            if operator == ">" and cell_num > val_num:
                                filtered.append(r)
                            elif operator == "<" and cell_num < val_num:
                                filtered.append(r)
                            elif operator == ">=" and cell_num >= val_num:
                                filtered.append(r)
                            elif operator == "<=" and cell_num <= val_num:
                                filtered.append(r)
                            elif operator == "==" and cell_num == val_num:
                                filtered.append(r)
                            elif operator == "!=" and cell_num != val_num:
                                filtered.append(r)
                        except (ValueError, TypeError):
                            if operator == "==" and str(cell).strip() == value:
                                filtered.append(r)
                            elif operator == "!=" and str(cell).strip() != value:
                                filtered.append(r)
                            elif operator == "contains" and value.lower() in str(cell).lower():
                                filtered.append(r)
                    data_rows = filtered
                    python_lines.append(f"df = df[df['{col}'] {operator} {repr(value)}]")

            elif action == "lowercase" and len(parts) >= 2:
                col = _sanitize_table_name(parts[1].strip())
                if col in headers:
                    col_idx = headers.index(col)
                    for r in data_rows:
                        if col_idx < len(r):
                            r[col_idx] = str(r[col_idx]).lower()
                    python_lines.append(f"df['{col}'] = df['{col}'].str.lower()")

            elif action == "to_numeric" and len(parts) >= 2:
                col = _sanitize_table_name(parts[1].strip())
                if col in headers:
                    col_idx = headers.index(col)
                    for r in data_rows:
                        if col_idx < len(r):
                            try:
                                r[col_idx] = str(float(str(r[col_idx]).replace(",", "")))
                            except (ValueError, TypeError):
                                r[col_idx] = ""
                    python_lines.append(f"df['{col}'] = pd.to_numeric(df['{col}'], errors='coerce')")

            elif action == "top" and len(parts) >= 3:
                try:
                    n = int(parts[1].strip())
                    col = _sanitize_table_name(parts[2].strip())
                    if col in headers:
                        col_idx = headers.index(col)
                        def _sort_key(r):
                            try:
                                return float(str(r[col_idx]).replace(",", ""))
                            except (ValueError, TypeError):
                                return 0.0
                        data_rows = sorted(data_rows, key=_sort_key, reverse=True)[:n]
                        python_lines.append(f"df = df.nlargest({n}, '{col}')")
                except (ValueError, IndexError):
                    pass

            elif action == "date_extract" and len(parts) >= 3:
                col = _sanitize_table_name(parts[1].strip())
                part_name = parts[2].strip().lower()
                if col in headers:
                    col_idx = headers.index(col)
                    new_col = f"{col}_{part_name}"
                    headers.append(new_col)
                    for r in data_rows:
                        cell = r[col_idx] if col_idx < len(r) else ""
                        extracted = ""
                        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]:
                            try:
                                dt = datetime.strptime(str(cell).strip(), fmt)
                                if part_name == "year":
                                    extracted = str(dt.year)
                                elif part_name == "month":
                                    extracted = str(dt.month)
                                elif part_name == "day":
                                    extracted = str(dt.day)
                                elif part_name == "weekday":
                                    extracted = str(dt.weekday())
                                break
                            except (ValueError, TypeError):
                                continue
                        r.append(extracted)
                    python_lines.append(f"df['{new_col}'] = pd.to_datetime(df['{col}']).dt.{part_name}")

            elif action == "aggregate" and len(parts) >= 4:
                group_col = _sanitize_table_name(parts[1].strip())
                agg_col = _sanitize_table_name(parts[2].strip())
                agg_func = parts[3].strip().lower()
                if group_col in headers and agg_col in headers:
                    g_idx = headers.index(group_col)
                    a_idx = headers.index(agg_col)
                    groups: Dict[str, List[float]] = {}
                    for r in data_rows:
                        key = r[g_idx] if g_idx < len(r) else ""
                        try:
                            val = float(str(r[a_idx]).replace(",", "")) if a_idx < len(r) else 0
                        except (ValueError, TypeError):
                            val = 0
                        groups.setdefault(key, []).append(val)
                    new_rows = []
                    for k, vals in groups.items():
                        if agg_func == "sum":
                            result = sum(vals)
                        elif agg_func == "avg":
                            result = sum(vals) / len(vals)
                        elif agg_func == "count":
                            result = len(vals)
                        elif agg_func == "min":
                            result = min(vals)
                        elif agg_func == "max":
                            result = max(vals)
                        else:
                            result = sum(vals)
                        new_rows.append([k, str(result)])
                    headers = [group_col, agg_col]
                    data_rows = new_rows
                    python_lines.append(f"df = df.groupby('{group_col}')['{agg_col}'].{agg_func}().reset_index()")

        output_csv_lines = [",".join(headers)]
        for row in data_rows[:MAX_PREVIEW_ROWS]:
            output_csv_lines.append(",".join(str(v) for v in row[:len(headers)]))

        python_code = "\n".join(python_lines)
        result = f"""# 🔄 DataPocket — Transformation Result

## Operations Applied
{chr(10).join(f"- `{op}`" for op in params.operations)}

## Transformed Data ({len(data_rows)} rows)
```csv
{chr(10).join(output_csv_lines)}
```

## Python Code (reproducible)
```python
{python_code}

# Save result
df.to_csv('transformed.csv', index=False)
```
"""
        if params.output_table_name:
            table = _sanitize_table_name(params.output_table_name)
            create_sql = _generate_create_table_sql(table, headers, data_rows)
            insert_sql = _generate_insert_sql(table, headers, data_rows)
            result += f"""
## SQL — Load to PostgreSQL
```sql
{create_sql}

{insert_sql}
```
"""
        result += "\n> Next step: Use `datapocket_generate_dashboard` to visualize this data."
        return result

    except (ParseError, TransformError) as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}"
    except Exception as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}. Sugerencia: Verifica que las operaciones usen el formato correcto con '|' como separador."


# ─── Tool 3: Generate Mobile-First Dashboard ───

class DashboardChartInput(BaseModel):
    """Definition for a single chart element in the dashboard."""
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")

    type: ChartType = Field(..., description="Chart type: bar, line, pie, kpi, table, area, donut")
    title: str = Field(..., description="Chart title", min_length=1, max_length=100)
    labels: Optional[List[str]] = Field(default=None, description="X-axis labels (for bar/line/pie/area/donut)")
    datasets: Optional[List[Dict[str, Any]]] = Field(
        default=None,
        description="Array of {label: str, data: [numbers]} for chart datasets"
    )
    value: Optional[str] = Field(default=None, description="KPI display value (e.g. '$12,500' or '1.2K')")
    subtitle: Optional[str] = Field(default=None, description="KPI subtitle (e.g. 'Promedio mensual')")
    trend: Optional[str] = Field(default=None, description="KPI trend indicator (e.g. '↑ 12.5%')")
    headers: Optional[List[str]] = Field(default=None, description="Table column headers")
    rows: Optional[List[List[str]]] = Field(default=None, description="Table rows (max 20 shown)")


class DashboardInput(BaseModel):
    """Input for dashboard generation."""
    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    title: str = Field(..., description="Dashboard title", min_length=1, max_length=100)
    charts: List[DashboardChartInput] = Field(..., description="List of charts/KPIs/tables to include", min_length=1, max_length=12)
    theme: str = Field(default="dark", description="Dashboard theme: 'dark' or 'light'")
    filename: Optional[str] = Field(default=None, description="Output HTML filename (without extension)")
    data_timestamp: Optional[str] = Field(default=None, description="Date label for the data period (e.g. 'Q1 2026')")


@mcp.tool(
    name="datapocket_generate_dashboard",
    annotations={
        "title": "Generate Mobile-First Dashboard",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def datapocket_generate_dashboard(params: DashboardInput) -> str:
    """Generate a mobile-first HTML dashboard with Chart.js visualizations.

    Creates a self-contained HTML file optimized for 320-480px mobile screens with:
    - KPI cards auto-arranged in 1 or 2-column grid
    - Interactive charts (bar, line, pie, area, donut)
    - Data tables with horizontal scroll
    - Dark/light theme with WCAG AA contrast
    - PWA-like bottom navigation bar
    - Separate data timestamp vs generation timestamp in header

    Args:
        params (DashboardInput): Validated input containing:
            - title (str): Dashboard title
            - charts (List[DashboardChartInput]): Chart/KPI/table definitions
            - theme (str): 'dark' or 'light'
            - filename (Optional[str]): Output filename (without .html)
            - data_timestamp (Optional[str]): Data period label

    Returns:
        str: Complete HTML dashboard code + hosting instructions.
    """
    try:
        charts_data = [chart.model_dump() for chart in params.charts]
        html = _generate_dashboard_html(params.title, charts_data, params.theme, params.data_timestamp)

        filename = params.filename or _sanitize_table_name(params.title)
        filepath = f"{filename}.html"

        result = f"""# 📱 DataPocket — Dashboard Generated

## {params.title}
- **Charts/KPIs**: {len(params.charts)}
- **Theme**: {params.theme}
- **File**: `{filepath}`

## How to Use
1. **Save** the HTML below to `{filepath}`
2. **Open** on your phone's browser or share via WhatsApp/Telegram
3. **Optional**: Host on GitHub Pages for a permanent URL

## Hosting on GitHub Pages (free)
```bash
git init datapocket-dashboards && cd datapocket-dashboards
# Save the HTML file here
git add . && git commit -m "Dashboard: {params.title}"
git push origin main
# Enable Pages in repo Settings → Pages → main branch
```

## Dashboard HTML
```html
{html}
```

---
> 📱 Tip: Add to phone home screen for app-like experience.
> 🔄 For live data, use `datapocket_query_to_dashboard`.
"""
        return result

    except Exception as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}. Sugerencia: Verifica que los charts tengan el formato correcto."


# ─── Tool 4: SQL Query to Dashboard ───

class QueryToDashboardInput(BaseModel):
    """Input for converting SQL query results to a dashboard."""
    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    query_results_csv: str = Field(..., description="Query results as CSV (paste from psql, DBeaver, or any SQL client)")
    dashboard_title: str = Field(..., description="Title for the generated dashboard", min_length=1, max_length=100)
    chart_suggestions: Optional[str] = Field(
        default=None,
        description="Hint about what charts to create. E.g., 'bar chart of sales by month, KPI for total revenue'"
    )
    theme: str = Field(default="dark", description="Dashboard theme: 'dark' or 'light'")


@mcp.tool(
    name="datapocket_query_to_dashboard",
    annotations={
        "title": "SQL Results → Mobile Dashboard",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def datapocket_query_to_dashboard(params: QueryToDashboardInput) -> str:
    """Convert SQL query results (CSV) into a mobile-first dashboard automatically.

    Analyzes the result structure, classifies columns by type, and generates
    appropriate KPI cards, charts, and a detail table — all in one HTML file.

    Args:
        params (QueryToDashboardInput): Validated input containing:
            - query_results_csv (str): Query results as CSV
            - dashboard_title (str): Dashboard title
            - chart_suggestions (Optional[str]): Visualization hints
            - theme (str): 'dark' or 'light'

    Returns:
        str: Column analysis + recommended dashboard configuration + HTML.
    """
    try:
        headers, data_rows = _parse_csv_data(params.query_results_csv)

        numeric_cols = []
        categorical_cols = []
        date_cols = []

        for i, h in enumerate(headers):
            col_values = [row[i] if i < len(row) else "" for row in data_rows]
            pg_type = _infer_pg_type(col_values)
            if pg_type in ("BIGINT", "NUMERIC"):
                numeric_cols.append(h)
            elif pg_type in ("DATE", "TIMESTAMP"):
                date_cols.append(h)
            else:
                categorical_cols.append(h)

        suggestions = []

        for col in numeric_cols[:4]:
            col_idx = headers.index(col)
            values = []
            for row in data_rows:
                try:
                    values.append(float(str(row[col_idx]).replace(",", "")))
                except (ValueError, IndexError):
                    pass
            if values:
                total = sum(values)
                avg = total / len(values)
                suggestions.append({
                    "type": "kpi",
                    "title": col.replace("_", " ").title(),
                    "value": _smart_format(total),
                    "subtitle": f"Avg: {_smart_format(avg)}",
                    "trend": ""
                })

        if categorical_cols and numeric_cols:
            cat_col = categorical_cols[0]
            num_col = numeric_cols[0]
            cat_idx = headers.index(cat_col)
            num_idx = headers.index(num_col)

            labels = [str(row[cat_idx]) if cat_idx < len(row) else "" for row in data_rows[:20]]
            data_vals = []
            for row in data_rows[:20]:
                try:
                    data_vals.append(float(str(row[num_idx]).replace(",", "")))
                except (ValueError, IndexError):
                    data_vals.append(0)

            chart_type = "line" if date_cols and cat_col in date_cols else "bar"
            suggestions.append({
                "type": chart_type,
                "title": f"{num_col.replace('_', ' ').title()} by {cat_col.replace('_', ' ').title()}",
                "labels": labels,
                "datasets": [{"label": num_col.replace("_", " ").title(), "data": data_vals}]
            })

        suggestions.append({
            "type": "table",
            "title": "Detail Data",
            "headers": [h.replace("_", " ").title() for h in headers],
            "rows": [[str(v)[:30] for v in row[:len(headers)]] for row in data_rows[:15]]
        })

        html = _generate_dashboard_html(params.dashboard_title, suggestions, params.theme)
        filename = _sanitize_table_name(params.dashboard_title)

        result = f"""# 📱 DataPocket — Auto-Dashboard from Query Results

## Data Analysis
- **Rows**: {len(data_rows)}
- **Numeric columns**: {', '.join(numeric_cols) or 'None'}
- **Categorical columns**: {', '.join(categorical_cols) or 'None'}
- **Date columns**: {', '.join(date_cols) or 'None'}

## Auto-Generated Charts
{chr(10).join(f"- **{s['type'].upper()}**: {s['title']}" for s in suggestions)}

## Dashboard HTML
Save as `{filename}.html` and open on your phone:

```html
{html}
```

---
> 💡 Customize: Use `datapocket_generate_dashboard` with specific chart configs for full control.
"""
        return result

    except (ParseError, ValidationError) as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}"
    except Exception as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}. Sugerencia: Verifica que el CSV tenga headers en la primera fila."


# ─── Tool 5: Power BI Connection ───

class PowerBIConnectionInput(BaseModel):
    """Input for generating Power BI connection setup."""
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")

    table_names: List[str] = Field(..., description="List of PostgreSQL table names to connect", min_length=1, max_length=20)
    schema_name: str = Field(default=DEFAULT_SCHEMA, description="PostgreSQL schema")
    pg_host: str = Field(default="localhost", description="PostgreSQL host")
    pg_port: str = Field(default="5432", description="PostgreSQL port")
    pg_database: str = Field(default="datapocket", description="PostgreSQL database name")


@mcp.tool(
    name="datapocket_powerbi_setup",
    annotations={
        "title": "Generate Power BI Connection",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def datapocket_powerbi_setup(params: PowerBIConnectionInput) -> str:
    """Generate Power BI Desktop connection setup for PostgreSQL tables.

    Creates step-by-step instructions and M (Power Query) code to connect
    Power BI Desktop to PostgreSQL tables — zero licensing cost.

    Args:
        params (PowerBIConnectionInput): Validated input containing:
            - table_names (List[str]): Tables to connect
            - schema_name (str): PostgreSQL schema
            - pg_host (str): Database host
            - pg_port (str): Database port
            - pg_database (str): Database name

    Returns:
        str: Complete Power BI setup guide with M code and starter DAX measures.
    """
    m_queries = []
    for table in params.table_names:
        safe_table = _sanitize_table_name(table)
        m_query = f'''let
    Source = PostgreSQL.Database("{params.pg_host}:{params.pg_port}", "{params.pg_database}"),
    {params.schema_name}_{safe_table} = Source{{[Schema="{params.schema_name}", Item="{safe_table}"]}}[Data]
in
    {params.schema_name}_{safe_table}'''
        m_queries.append((safe_table, m_query))

    queries_section = ""
    for table, mq in m_queries:
        queries_section += f"""
### Table: `{table}`
```powerquery
{mq}
```
"""

    return f"""# ⚡ DataPocket — Power BI Setup

## Prerequisites
1. **Power BI Desktop** (free): [Download](https://powerbi.microsoft.com/desktop/)
2. **Npgsql driver** for PostgreSQL: [Download](https://github.com/npgsql/npgsql/releases)
3. PostgreSQL running at `{params.pg_host}:{params.pg_port}`

## Step-by-Step Connection
1. Open Power BI Desktop
2. Click **Get Data** → **Database** → **PostgreSQL database**
3. Enter Server: `{params.pg_host}:{params.pg_port}` and Database: `{params.pg_database}`
4. Select **DirectQuery** for real-time or **Import** for snapshots
5. Select your tables and click **Load**

## Power Query M Code (Advanced)
Use in **Advanced Editor** (Home → Transform Data → Advanced Editor):

{queries_section}

## DAX Measures (Starter Kit)
```dax
// Total Revenue
Total Revenue = SUM('{m_queries[0][0]}'[revenue])

// Month-over-Month Growth
MoM Growth =
VAR CurrentMonth = [Total Revenue]
VAR PreviousMonth = CALCULATE([Total Revenue], DATEADD('Calendar'[Date], -1, MONTH))
RETURN DIVIDE(CurrentMonth - PreviousMonth, PreviousMonth, 0)

// Running Total
Running Total = CALCULATE([Total Revenue], FILTER(ALL('Calendar'[Date]), 'Calendar'[Date] <= MAX('Calendar'[Date])))
```

## Mobile Layout (Power BI)
1. Go to **View** → **Mobile layout**
2. Drag visuals to the phone canvas
3. Publish to Power BI Service (free with work/school account)
4. Access via **Power BI Mobile app** on your phone

---
> 🎯 Free ETL stack: PostgreSQL + Python + Power BI = $0 licensing.
> Compare: Tableau ($75/user/mo), Monday.com Pro ($16/user/mo).
"""


# ─── Tool 6: Export Data ───

class ExportInput(BaseModel):
    """Input for exporting data in multiple formats."""
    model_config = ConfigDict(str_strip_whitespace=True, validate_assignment=True, extra="forbid")

    data: str = Field(..., description="CSV data to export (with headers in first row)")
    format: ExportFormat = Field(..., description="Target export format: csv, json, jsonl, sql, markdown")
    table_name: Optional[str] = Field(default=None, description="Table name for SQL INSERT format")


@mcp.tool(
    name="datapocket_export",
    annotations={
        "title": "Export Data in Multiple Formats",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def datapocket_export(params: ExportInput) -> str:
    """Convert CSV data to CSV, JSON array, JSONL, SQL INSERTs, or Markdown table.

    Useful for converting between data formats, preparing data for APIs,
    generating SQL migration scripts, or creating Markdown reports.

    Args:
        params (ExportInput): Validated input containing:
            - data (str): Source CSV data with headers
            - format (ExportFormat): Target format (csv, json, jsonl, sql, markdown)
            - table_name (Optional[str]): Table name for SQL format

    Returns:
        str: Converted data in the requested format.
    """
    try:
        headers, data_rows = _parse_csv_data(params.data)

        if params.format == ExportFormat.CSV:
            lines = [",".join(headers)]
            for row in data_rows:
                lines.append(",".join(f'"{str(v)}"' if "," in str(v) else str(v) for v in row[:len(headers)]))
            output = "\n".join(lines)
            return f"## Exported CSV\n```csv\n{output}\n```"

        elif params.format == ExportFormat.JSON:
            records = []
            for row in data_rows:
                obj = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
                records.append(obj)
            output = json.dumps(records, ensure_ascii=False, indent=2)
            return f"## Exported JSON\n```json\n{output}\n```"

        elif params.format == ExportFormat.JSONL:
            lines = []
            for row in data_rows:
                obj = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
                lines.append(json.dumps(obj, ensure_ascii=False))
            output = "\n".join(lines)
            return f"## Exported JSONL\n```\n{output}\n```"

        elif params.format == ExportFormat.SQL:
            table = _sanitize_table_name(params.table_name or "exported_data")
            create_sql = _generate_create_table_sql(table, headers, data_rows)
            insert_sql = _generate_insert_sql(table, headers, data_rows)
            return f"## Exported SQL\n```sql\n{create_sql}\n\n{insert_sql}\n```"

        elif params.format == ExportFormat.MARKDOWN_TABLE:
            header_row = "| " + " | ".join(headers) + " |"
            sep_row = "| " + " | ".join(["---"] * len(headers)) + " |"
            data_rows_md = []
            for row in data_rows:
                data_rows_md.append("| " + " | ".join(str(v)[:50] for v in row[:len(headers)]) + " |")
            output = "\n".join([header_row, sep_row] + data_rows_md)
            return f"## Exported Markdown Table\n\n{output}"

        else:
            return f"❌ Error (ValidationError): Formato '{params.format}' no soportado. Sugerencia: Usa csv, json, jsonl, sql, o markdown."

    except (ParseError, ValidationError) as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}"
    except Exception as e:
        return f"❌ Error ({type(e).__name__}): {str(e)}. Sugerencia: Verifica que el CSV de entrada tenga headers válidos."


# ──────────────────────────────────────────────
# Entry Point
# ──────────────────────────────────────────────

def main() -> None:
    """Run the DataPocket MCP server."""
    mcp.run()


if __name__ == "__main__":
    main()
